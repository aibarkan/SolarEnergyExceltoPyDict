import openpyxl

# import xl_functions


# CONSTANTS
LOCALJXB = r"C:\Users\Jason\petrus\coding\repos\localstorage/BCA - Vendor1 proposal Baldwinsville (10-03-17) Lite - slh v2.xlsm"
AXB_FILE = "/Users/alexandrabarkan/Desktop/SOM/BCA - Vendor1 proposal Baldwinsville (10-03-17) Lite - slh v2.xlsx"
EXCEL_FILE = LOCALJXB
PROJROB_SHEETNAME = "4.2-4.4_Benefits"
PROJROB_PROJINPSHEET = "Inputs - Project Specific"
PROJROB_FIXEDSHEET = "Inputs - Fixed"
MAX_EXCEL_BOUNDARY = 1000
MAX_LIST_LENGTH = 100

# Make a Input Class
'''
class Table_Index:
    def __init__(self,my_workbook,sheet, beginning_col_str, ending_col_str):
        self.wb = my_workbook
        self.start_col_loc =()
        self.end_col_loc = ()
        self.end_row_loc = []
        self.sheet = sheet
        self.start_col_str = beginning_col_str
        self.end_col_str = ending_col_str
    def xl_get_cell_loc_containing_text(self, sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> tuple[
        int, int, int]:
        retval = -1
        print('hi')
        for row in range(start_row, start_row+max_boundary):
            for col in range(start_col, start_col+max_boundary):
                cell_val = sheet.cell(row, col)
                if str(cell_val.value) == str_val:
                    retval = 0
                    return row, col, retval

'''
def main():
    get_total_dict()


def get_total_dict():
    my_workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    get_project_specific_assumptions(my_workbook)
    # my_input_dict = make_input_dict(my_workbook)


# PUT IN LIBRBARY
def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> \
        tuple[
            int, int, int]:
    retval = -1
    print('hi')
    for row in range(start_row, start_row + max_boundary):
        for col in range(start_col, start_col + max_boundary):
            cell_val = sheet.cell(row, col)
            if str(cell_val.value) == str_val:
                retval = 0
                return row, col, retval


def xl_get_first_empty_cell(sheet, start_row: int = 1, start_col: int = 1, max_listlength: int = MAX_LIST_LENGTH):
    retval = -1
    for row in range(start_row, start_row + max_listlength):
        print(f"Testing cell in row: {row}, col: {start_col}")
        cell_val = sheet.cell(row, start_col)
        print(f">> Current cell val: {str(cell_val.value)} ")
        if cell_val.value in [None,'',""]:
            retval = 0
            return row, start_col, retval
    return 0, 0, retval
    #return row, start_col

def get_project_specific_assumptions(my_workbook):
    proj_assump_str = "Project-Specific Assumptions"
    proj_assump_notes_str = "Notes"
    proj_spec_inputs_sheetname = "Inputs - Project Specific"
    proj_spec_inputs_sheet = my_workbook[proj_spec_inputs_sheetname]
    # get_table_index(my_workbook, proj_spec_inputs_sheet,proj_assump_str,proj_assump_notes_str)
    # start location of proj assump table
    # project_specific_assump_loc = xl_get_cell_loc_containing_text(my_sheet, proj_assump_str)
    project_specific_assump_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_str)
    print(project_specific_assump_loc)
    project_specific_assump_notes_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_notes_str)
    print(project_specific_assump_notes_loc)
    project_specific_assump_end_row_loc = xl_get_first_empty_cell(proj_spec_inputs_sheet, max_listlength=MAX_LIST_LENGTH,
                                                                  start_col=project_specific_assump_loc[1],
                                                                  start_row=project_specific_assump_loc[0])
    print(project_specific_assump_end_row_loc)


main()
