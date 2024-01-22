import openpyxl
from collections import OrderedDict
#CONSTANTS
EXCEL_FILE = r"C:\Users\abarkan\Documents\GirlsontheRun\KPMG Report 1.1.2019 to 11.18.2022 donor data.xlsx"
MAX_EXCEL_BOUNDARY = 25000


def get_job_configuration_values() -> dict:
    inputs_dict = {
        "Account Name":
            {"Cols": ["Account Name", "Opportunity Name", "Amount","Close Date","Lead Source","Solicited By",\
                      "Billing Zip/Postal Code","Account Codes","Class Codes","Total Gifts This Year",\
                      "Total Gifts Last Year","Total Gifts"],
             "Adjust": [0, 0],
             "Sheetname": "KPMG Report 6 2022",
             "Inputdata": {}
             }
    }
    return inputs_dict


def main():
    my_workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    my_input_dict = get_job_configuration_values()
    input_dicts = get_input_data(my_input_dict, my_workbook)
    return input_dicts


def get_input_data(inputsdict, my_workbook)-> dict:
    for dict in inputsdict.keys():
        tables_dict_sheetname = inputsdict[dict]["Sheetname"]
        sheet = my_workbook[tables_dict_sheetname]
        anchor_table_loc = xl_get_cell_loc_containing_text(sheet, dict)
        print(anchor_table_loc)
        end_table_loc = xl_get_first_empty_cell(sheet, MAX_EXCEL_BOUNDARY, anchor_table_loc[0], anchor_table_loc[1])
        adjust = inputsdict[dict]["Adjust"]
        col_names_dict = xl_get_all_str_in_row_to_dict(sheet, start_row=anchor_table_loc[0]+adjust[0],
                                                       start_col=anchor_table_loc[1]+1)
        input_dict = make_input_table_dict(sheet, col_names_dict, start_row = anchor_table_loc[0]+adjust[1],
                                           end_row= end_table_loc[0], start_col= anchor_table_loc[1])
        inputsdict[dict]["Inputdata"].update(input_dict)
    return inputsdict

def make_input_table_dict(sheet, colsdict, start_row: int = 1, end_row: int = MAX_EXCEL_BOUNDARY,
                          start_col: int = 1):
    inputvalsdict={}
    for row in range(start_row, end_row):
        factornamecell=sheet.cell(row, start_col)
        factorname=str(factornamecell.value)
        colvals={}
        for colname in colsdict.keys():
            colvalcell=sheet.cell(row, colsdict.get(colname))
            colvalval = str(colvalcell.value)
            colvals[colname]=colvalval
        inputvalsdict[factorname]=colvals
    return inputvalsdict

#Excel Functions
def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary = MAX_EXCEL_BOUNDARY) -> \
        tuple[int, int]:
    for row in range(start_row, start_row+1):
        for col in range(start_col, start_col+1):
            cell_val = sheet.cell(row, col)
            if str(cell_val.value) == str_val:
                print(cell_val)
                return row, col

def xl_get_first_empty_cell(sheet, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1) -> \
        tuple[int, int]:
    retval = -1
    for row in range(start_row, start_row + end_index):
        cell_val = sheet.cell(row, start_col)
        if cell_val.value in [None, '', ""]:
            retval = 0
            return row, start_col
    return row, retval

def xl_get_all_str_in_row_to_dict(sheet, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1):
    my_dict = {}
    for col in range(start_col, start_col + end_index):
        retval = -1
        cell_val = sheet.cell(start_row, col)
        if cell_val.value not in [None, '', ""]:
            retval = 0
            my_dict.update({cell_val.value: col})
    return my_dict





main()
