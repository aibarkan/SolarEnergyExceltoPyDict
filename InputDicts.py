import openpyxl
from collections import OrderedDict
#CONSTANTS
EXCEL_FILE = "/Users/alexandrabarkan/Desktop/SOM/BCA - Vendor1 proposal Baldwinsville wc Lite - slh v2.xlsm"
MAX_EXCEL_BOUNDARY = 1000


def get_job_configuration_values() -> dict:
    inputs_dict = {
        "Project-Specific Assumptions":
            {"Cols": ["Value", "Unit", "Notes"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Project Specific",
             "Inputdata": {}
             },
        "Forecasted Load Growth / Need":
            {"Cols": ["Growth", "Unit", "Notes"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Project Specific",
             "Inputdata": {}
             },
        "Technology Assumptions":
            {"Cols": ["Applicable", "Year", "Value", "Unit", "Notes"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Project Specific",
             "Inputdata": {}
             },
        "General Assumptions":
            {"Cols": ["Value", "Unit", "Source"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Fixed",
             "Inputdata": {}
             },
        "Cost of CO2, SOx and NOx":
            {"Cols": ["SOx", "NOx", "CO2", "Unit", "Source"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Fixed",
             "Inputdata": {}
             },
        "Air Emissions":
            {
                "Adjust": [1, 2],
                "Sheetname": "Inputs - Fixed",
                "Inputdata": {}
            },
        "Avoided Generation Capacity Costs (AGCC), Upstate Load Zones A-F  - Provided by DPS Staff 1/21/2016 ":
            {"Cols": ["AGCC Value", "Unit", "Value", "Source"],
             "Adjust": [0, 1],
             "Sheetname": "Inputs - Fixed",
             "Inputdata": {}
             },
        "Avoided LBMP - CARIS 2 Base Case Annual Average LBMP ($/MWh) by Load Zone - posted 7/14/2016":
            {
                "Adjust": [1, 2],
                "Sheetname": "Inputs - Fixed",
                "Inputdata": {}
            },
        "Utility System Average Marginal Cost of Service ($/kW-yr)":
            {
                "Adjust": [1, 2],
                "Sheetname": "Inputs - Fixed",
                "Inputdata": {}
            }
    }
    return inputs_dict


def get_input_dicts():
    my_workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    my_input_dict = get_job_configuration_values()
    input_dicts = get_input_data(my_input_dict, my_workbook)
    return input_dicts


def get_input_data(inputsdict, my_workbook)-> dict:
    for dict in inputsdict.keys():
        tables_dict_sheetname = inputsdict[dict]["Sheetname"]
        sheet = my_workbook[tables_dict_sheetname]
        anchor_table_loc = xl_get_cell_loc_containing_text(sheet, dict)
        end_table_loc = xl_get_first_empty_cell(sheet, MAX_EXCEL_BOUNDARY, anchor_table_loc[0], anchor_table_loc[1])
        adjust = inputsdict[dict]["Adjust"]
        col_names_dict = xl_get_all_str_in_row_to_dict(sheet, start_row=anchor_table_loc[0]+adjust[0],
                                                           start_col=anchor_table_loc[1]+1)
        input_dict = make_input_table_dict(sheet, col_names_dict, start_row = anchor_table_loc[0]+adjust[1],
                                           end_row= end_table_loc[0], start_col= anchor_table_loc[1])
        inputsdict[dict]["Inputdata"].update(input_dict)
    return inputsdict

def make_input_table_dict(sheet, colsdict, start_row: int = 1, end_row: int = MAX_EXCEL_BOUNDARY,
                          start_col: int = 1):
    inputvalsdict={}
    for row in range(start_row, end_row):
        factornamecell=sheet.cell(row, start_col)
        factorname=str(factornamecell.value)
        colvals={}
        for colname in colsdict.keys():
            colvalcell=sheet.cell(row, colsdict.get(colname))
            colvalval = str(colvalcell.value)
            colvals[colname]=colvalval
        inputvalsdict[factorname]=colvals
    return inputvalsdict

#Excel Functions
def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary = MAX_EXCEL_BOUNDARY) -> tuple:
    for row in range(start_row, start_row + max_boundary):
        for col in range(start_col, start_col + max_boundary):
            cell_val = sheet.cell(row, col)
            if str(cell_val.value) == str_val:
                return row, col

def xl_get_first_empty_cell(sheet, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1) -> tuple:
    retval = -1
    for row in range(start_row, start_row + end_index):
        cell_val = sheet.cell(row, start_col)
        if cell_val.value in [None, '', ""]:
            retval = 0
            return row, start_col
    return row, retval

def xl_get_all_str_in_row_to_dict(sheet, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1):
    my_dict = {}
    for col in range(start_col, start_col + end_index):
        retval = -1
        cell_val = sheet.cell(start_row, col)
        if cell_val.value not in [None, '', ""]:
            retval = 0
            my_dict.update({cell_val.value: col})
    return my_dict


