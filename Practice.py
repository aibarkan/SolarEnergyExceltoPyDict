import openpyxl

# import xl_functions


# CONSTANTS
EXCEL_FILE = "/Users/alexandrabarkan/Desktop/SOM/BCA - Vendor1 proposal Baldwinsville (10-03-17) Lite - slh v2.xlsx"
PROJROB_SHEETNAME = "4.2-4.4_Benefits"
PROJROB_PROJINPSHEET = "Inputs - Project Specific"
PROJROB_FIXEDSHEET = "Inputs - Fixed"
MAX_EXCEL_BOUNDARY = 1000

# Make a Input Class
'''
class Table_Index:
    def __init__(self,my_workbook,sheet, beginning_col_str, ending_col_str):
        self.wb = my_workbook
        self.start_col_loc =()
        self.end_col_loc = ()
        self.end_row_loc = []
        self.sheet = sheet
        self.start_col_str = beginning_col_str
        self.end_col_str = ending_col_str
    def xl_get_cell_loc_containing_text(self, sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> tuple[
        int, int, int]:
        retval = -1
        print('hi')
        for row in range(start_row, start_row+max_boundary):
            for col in range(start_col, start_col+max_boundary):
                cell_val = sheet.cell(row, col)
                if str(cell_val.value) == str_val:
                    retval = 0
                    return row, col, retval

'''


def main():
    get_total_dict()


def get_total_dict():
    my_workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    get_project_specific_assumptions(my_workbook)
    # my_input_dict = make_input_dict(my_workbook)


# PUT IN LIBRBARY
def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> \
        tuple[int, int, int]:
    retval = -1
    for row in range(start_row, start_row + max_boundary):
        for col in range(start_col, start_col + max_boundary):
            cell_val = sheet.cell(row, col)
            if str(cell_val.value) == str_val:
                retval = 0
                return row, col, retval


def xl_get_first_empty_cell(sheet, end_index: int = 1000, start_row: int = 1, start_col: int = 1):
    for row in range(start_row, start_row + end_index):
        cell_val = sheet.cell(row, start_col)
        print(cell_val.value)
        if cell_val.value in [None, '', ""]:
            print(cell_val.value)
            return row, start_col
    return 0,0


def ref():
    CONSTDICTOFTABLES={"Project-Specific Assumptions":["Value","Unit","Notes"], "Technology_Assumptions":["Value","Unit","Notes"]}
    for tables in CONSTDICTOFTABLES:
        make_input_table_dict()

CONSTDICTOFTABLES={"Project-Specific Assumptions":["Value","Unit","Notes"], "Technology_Assumptions":["Value","Unit","Notes"]}

def make_input_table_dict(sheet, list_of_tablenames:str, start_row: int = 1, end_row: int = 1000, start_col: int = 1, end_col: int = 1):
    colsdict={"Value":6,"Unit":7,"Notes":8}
    inputvalsdict={}
    for row in range(start_row, start_row+end_row+2):
        factornamecell=sheet.cell(row, start_col)
        factorname=str(factornamecell.value)
        colvals={}
        for colname in colsdict.keys():
            colvalcell=sheet.cell(row, colsdict[colname].value)
            colvalval = str(colvalcell.value)
            colvals[colname]=colvalval
        inputvalsdict[factorname]=colvals
    return inputvalsdict




def make_outer_dict(sheet, dict, list, start_row: int = 1, end_row: int = 1000, start_col: int = 1, end_col: int = 1):
    for row in range(start_row, start_row+end_row+2):
        for col in range(start_col, end_col):
            cell_val = sheet.cell(row,col)
            #list.append(cell_val.value)
            dict[cell_val.value] = {}
            print(dict)
def make_inner_dict_keys(sheet,dict_keys, dict, start_row: int = 1, end_row: int = 1000, start_col: int = 1, end_col: int = 1):
    for row in range(1, start_row+1):
        print(start_row)
        for col in range(start_col, start_col + end_col):
            cell_val = sheet.cell(row, col)
            print(cell_val)
            print(cell_val.value)
            if cell_val.value:
                print('h')
                dict_keys.append(cell_val.value)
                print(dict_keys)
                dict = dict[][cell_val.value]
                print(dict)


                # iterate through first row = outer dict key
    # iterate through rest of dict =
    '''   for row in range(start_row, start_row+max_boundary):
        for col in range(start_col, start_col + max_boundary):
            cell_val = sheet.cell(row,col)
            print(cell_val.value)'''


def get_project_specific_assumptions(my_workbook):
    proj_assump_str = "Project-Specific Assumptions"
    proj_assump_notes_str = "Notes"
    proj_spec_inputs_sheetname = "Inputs - Project Specific"
    proj_spec_inputs_sheet = my_workbook[proj_spec_inputs_sheetname]
    # get_table_index(my_workbook, proj_spec_inputs_sheet,proj_assump_str,proj_assump_notes_str)
    # start location of proj assump table
    # project_specific_assump_loc = xl_get_cell_loc_containing_text(my_sheet, proj_assump_str)
    project_specific_assump_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_str)
    print(project_specific_assump_loc)
    project_specific_assump_notes_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_notes_str)
    print(project_specific_assump_notes_loc)
    #project_spec_assumptions_dict[assump] = {}
    first_cell = project_specific_assump_loc[0]
    firstr_endcol = project_specific_assump_notes_loc[1]
    outer_keys = []
    #make_outer_dict(proj_spec_inputs_sheet, project_spec_assumptions_dict, outer_keys, start_col=first_cell, end_col=first_cell+2,start_row=first_cell, end_row=firstr_endcol)
    inner_keys = []
    #make_inner_dict_keys(proj_spec_inputs_sheet, inner_keys, dict=project_spec_assumptions_dict, start_row = first_cell, end_row= first_cell+2, start_col=first_cell+1, end_col=firstr_endcol)
    #project_spec_assumptions_dict[outer_keys].update(inner_keys)
    #print(project_spec_assumptions_dict)
    project_specific_assump_end_row_loc = xl_get_first_empty_cell(proj_spec_inputs_sheet, end_index=20, start_col=project_specific_assump_loc[1],start_row=project_specific_assump_loc[0])
    # start_col=project_specific_assump_loc[0],
    # start_row=project_specific_assump_loc[1])
    print(project_specific_assump_end_row_loc)


main()
