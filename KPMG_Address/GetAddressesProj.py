import openpyxl

# import xl_functions


# CONSTANTS
LOCALJXB = r"C:\Users\Jason\petrus\coding\repos\localstorage/BCA - Vendor1 proposal Baldwinsville (10-03-17) Lite - slh v2.xlsm"
AXB_FILE = "/Users/alexandrabarkan/Desktop/SOM/BCA - Vendor1 proposal Baldwinsville (10-03-17) Lite - slh v2.xlsx"
JXB_FILE = "/Users/jbarkan/repos/axios/KPMG_Address/KPMG_ReportFile.xlsx"
EXCEL_FILE = JXB_FILE
KPMG_SHEET = "KPMG Report 6 2022"
MAX_EXCEL_BOUNDARY = 1000
MAX_LIST_LENGTH = 100
AddressesCol = 8
StartRow = 2
EndRow = 10


def main():
    run_program()


def run_program():
    xlworkbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    addressesSheet = xlworkbook[KPMG_SHEET]
    zipcodes = get_all_zip_codes(addressesSheet, AddressesCol, StartRow)# get_project_specific_assumptions(my_workbook)
    scratch()

def get_all_zip_codes(sheet, start_col, start_row: int = 2) -> dict:
    rawzipcodes = []
    for row in range(2, 10):
        cell_val = sheet.cell(row, start_col).value
        if(len(cell_val) == 5):
            rawzipcodes.append(cell_val)
    unique_zipcodes = list(set(rawzipcodes))
    output_dict = populate_dict_of_zipcodes_with_addresses(unique_zipcodes)
    print(output_dict)
    return output_dict


def populate_dict_of_zipcodes_with_addresses(unique_zipcodes: list) -> dict:
    output_dict = {}
    for zipcode in unique_zipcodes:
        addr_dict = get_single_address_from_zip(zipcode)
        output_dict[zipcode] = addr_dict
    return output_dict

def get_single_address_from_zip(zipcode: str) -> dict:
    zipstring = zipcode_searcher(zipcode)
    single_addr_dict = zipstring_to_dict(zipstring)
    return single_addr_dict


def zipcode_searcher(zipcode: str) -> str:
    zipstring = "Tulsa, Oklahoma, " + zipcode + ", USA"
    return zipstring

def zipstring_to_dict(zipstring: str) -> dict:
    address_dict = {'City': '', 'State': '', 'Zip': '', 'Country': ''}
    lst = zipstring.split()
    address_dict['City'] = lst[0]
    address_dict['State'] = lst[1]
    address_dict['Zip'] = lst[2]
    address_dict['Country'] = lst[3]
    return address_dict

def scratch():
    print("Done")


# PUT IN LIBRBARY

#def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> tuple[int, int, int]:
#    retval = -1
#    print('hi')
#    for row in range(start_row, start_row + max_boundary):
#        for col in range(start_col, start_col + max_boundary):
#            cell_val = sheet.cell(row, col)
#            if str(cell_val.value) == str_val:
#                retval = 0
#                return row, col, retval




def xl_get_first_empty_cell(sheet, start_row: int = 1, start_col: int = 1, max_listlength: int = MAX_LIST_LENGTH):
    retval = -1
    for row in range(start_row, start_row + max_listlength):
        print(f"Testing cell in row: {row}, col: {start_col}")
        cell_val = sheet.cell(row, start_col)
        print(f">> Current cell val: {str(cell_val.value)} ")
        if cell_val.value in [None,'',""]:
            retval = 0
            return row, start_col, retval
    return 0, 0, retval
    #return row, start_col

def get_project_specific_assumptions(my_workbook):
    proj_assump_str = "Project-Specific Assumptions"
    proj_assump_notes_str = "Notes"
    proj_spec_inputs_sheetname = "Inputs - Project Specific"
    proj_spec_inputs_sheet = my_workbook[proj_spec_inputs_sheetname]
    # get_table_index(my_workbook, proj_spec_inputs_sheet,proj_assump_str,proj_assump_notes_str)
    # start location of proj assump table
    # project_specific_assump_loc = xl_get_cell_loc_containing_text(my_sheet, proj_assump_str)
    project_specific_assump_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_str)
    print(project_specific_assump_loc)
    project_specific_assump_notes_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_notes_str)
    print(project_specific_assump_notes_loc)
    project_specific_assump_end_row_loc = xl_get_first_empty_cell(proj_spec_inputs_sheet, max_listlength=MAX_LIST_LENGTH,
                                                                  start_col=project_specific_assump_loc[1],
                                                                  start_row=project_specific_assump_loc[0])
    print(project_specific_assump_end_row_loc)


main()
