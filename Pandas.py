import openpyxl
from collections import OrderedDict

# CONSTANTS
EXCEL_FILE = "/Users/alexandrabarkan/Desktop/SOM/BCA - Vendor1 proposal Baldwinsville wc Lite - slh v2.xlsm"
PROJROB_SHEETNAME = "4.2-4.4_Benefits"
PROJROB_PROJINPSHEET = "Inputs - Project Specific"
PROJROB_FIXEDSHEET = "Inputs - Fixed"
MAX_EXCEL_BOUNDARY = 1000
LISTOFINPUTTABLES=[]

def main():
    get_total_dict()

def get_total_dict():
    my_workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    get_project_input_dicts(my_workbook)

def get_project_input_dicts(my_workbook):
    LISTOFINPUTTABLES.append(get_project_specific_assumptions_dicts(my_workbook))


def get_job_configuration_values()-> dict:
    jobdict={"workbook_path":"\\users\alex\etcetc"}
    tablesdict={
        "Project-Specific Assumptions":
            {"Cols":["Value", "Unit", "Notes"],
             "Sheetname": "Inputs - Project Specific",
             "Inputdata":{}
             },
        "Forecasted Load Growth":
            {"Cols":["Growth", "Unit", "Notes"],
             "Sheetname": "Inputs - Project Specific",
             "Inputdata":{}
             },
    }
    jobdict["InputTablesConfig":tablesdict]

def main():
    jobdict=get_job_configuration_values()
    jobdict=get_input_data()
    cal_output

def get_input_data(jobdict:dict)-> dict:
    sheet =
    #for each table in jobdict.keys():




def get_project_specific_assumptions_dicts(my_workbook):
    #PROJ_SPEC_ASSUMP_CONSTANTS
    proj_assump_str = "Project-Specific Assumptions"
    proj_assump_notes_str = "Notes"
    proj_spec_inputs_sheetname = "Inputs - Project Specific"
    proj_spec_inputs_sheet = my_workbook[proj_spec_inputs_sheetname]
    forecasted_load_growth_str = "Forecasted Load Growth / Need"
    forecasted_load_growth_year_str = "Year"

    #Get location of proj_sepc_assump table
    '''
    project_specific_assump_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_str)

    #Get location of last cell of table
    project_specific_assump_notes_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_notes_str)

    #Get length of table
    project_specific_assump_end_row_loc = xl_get_first_empty_cell(proj_spec_inputs_sheet,
                                                                  start_col=project_specific_assump_loc[1],
                                                                  start_row=project_specific_assump_loc[0])
    project_specific_assump_keys_dict = {}
    project_specific_assump_keys_dict = xl_get_all_str_in_row_to_dict(proj_spec_inputs_sheet, project_specific_assump_keys_dict,
                                                                   start_row=project_specific_assump_loc[0],
                                                                   start_col=project_specific_assump_loc[0]+1)
    print(project_specific_assump_keys_dict)
    project_specific_assump_dict = make_input_table_dict(proj_spec_inputs_sheet, project_specific_assump_keys_dict,
                                                         start_row=project_specific_assump_loc[0] + 1,
                                                         end_row=project_specific_assump_end_row_loc[0],
                                                         start_col=project_specific_assump_loc[1])
    print(project_specific_assump_dict)
    '''
    #FORCASTED LOAD GROWTH
    forecasted_load_growth_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, forecasted_load_growth_str)

    forecasted_load_growth_notes_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet, proj_assump_notes_str,
                                                                       start_row=forecasted_load_growth_loc[0])
    print(forecasted_load_growth_notes_loc)
    forecasted_load_growth_year_loc = xl_get_cell_loc_containing_text(proj_spec_inputs_sheet,
                                                                      forecasted_load_growth_year_str,
                                                                      start_row=forecasted_load_growth_loc[0])
    forecasted_load_growth_end_row_loc = xl_get_first_empty_cell(proj_spec_inputs_sheet,
                                                                 start_col=forecasted_load_growth_year_loc[1],
                                                                 start_row=forecasted_load_growth_year_loc[0])
    forecasted_load_growth_keys_dict = OrderedDict()
    forecasted_load_growth_keys_dict = xl_get_all_str_in_row_to_dict(proj_spec_inputs_sheet,
                                                                    forecasted_load_growth_keys_dict,
                                                                    start_row=forecasted_load_growth_loc[0],
                                                                    start_col=forecasted_load_growth_loc[1],
                                                                    end_index=forecasted_load_growth_notes_loc[1])
    print(forecasted_load_growth_keys_dict)
    #start_row=project_specific_assump_loc[0])
    print(forecasted_load_growth_loc[1])
    forecasted_load_growth_dict = make_input_table_dict(proj_spec_inputs_sheet, key_name=True,
                                                        colsdict = forecasted_load_growth_keys_dict,
                                                         start_row=forecasted_load_growth_loc[0] + 1,
                                                        end_row=forecasted_load_growth_end_row_loc[0],
                                                         start_col=forecasted_load_growth_year_loc[1])
    return project_specific_assump_dict, forecasted_load_growth_dict





#Functions
def make_input_table_dict(sheet, colsdict, key_name = False, start_row: int = 1, end_row: int = MAX_EXCEL_BOUNDARY,
                          start_col: int = 1):

    inputvalsdict={}
    for row in range(start_row, end_row):
        factornamecell=sheet.cell(row, start_col)
        factorname=str(factornamecell.value)
        colvals={}
        if key_name == True:
            keyname = list(colsdict)[0]
            #factorname = keyname
            #inputvalsdict[keyname] = {}
            del colsdict[next(iter(colsdict))]
            key_name = False
            print(colsdict)
        else:
            for colname in colsdict.keys():
                colvalcell=sheet.cell(row, colsdict.get(colname))
                print(colsdict.get(colname))
                colvalval = str(colvalcell.value)
                colvals[colname]=colvalval
        inputvalsdict[factorname]=colvals
    print(inputvalsdict)
    return inputvalsdict

def xl_get_cell_loc_containing_text(sheet, str_val, start_col: int = 1, start_row: int = 1, max_boundary: int = 1000) -> \
        tuple[int, int, int]:
    retval = -1
    for row in range(start_row, start_row + max_boundary):
        for col in range(start_col, start_col + max_boundary):
            cell_val = sheet.cell(row, col)
            if str(cell_val.value) == str_val:
                retval = 0
                return row, col, retval


def xl_get_first_empty_cell(sheet, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1):
    for row in range(start_row, start_row + end_index):
        cell_val = sheet.cell(row, start_col)
        if cell_val.value in [None, '', ""]:
            return row, start_col


def xl_get_all_str_in_row_to_dict(sheet, my_dict, end_index: int = MAX_EXCEL_BOUNDARY, start_row: int = 1, start_col: int = 1):
    for col in range(start_col, start_col + end_index):
        cell_val = sheet.cell(start_row, col)
        if cell_val.value not in [None, '', ""]:
            my_dict.update({cell_val.value: col})
    return my_dict




main()